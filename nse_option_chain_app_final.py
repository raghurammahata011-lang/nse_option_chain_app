# nse_option_chain_app.py
# Professional, full-featured Streamlit app for NSE option chain analytics,
# watchlist, strategy signals, and backtesting (underlying via yfinance).
#
# Requirements:
# streamlit, pandas, numpy, plotly, scikit-learn, openpyxl, requests, kaleido, yfinance

import time
import random
import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.io as pio
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from sklearn.linear_model import LinearRegression, Ridge, Lasso, LogisticRegression
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score, accuracy_score, confusion_matrix, classification_report
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
import os
import concurrent.futures
import yfinance as yf
import warnings
warnings.filterwarnings('ignore')

# ---------------- CONFIG ----------------
SAVE_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "NSE_STOCK")
os.makedirs(SAVE_FOLDER, exist_ok=True)
INDICES = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"]
# tune this if necessary
THREAD_WORKERS = 6

# ---------------- PROFESSIONAL CSS ----------------
PRO_CSS = """
<style>
:root {
    --primary: #0b69ff;
    --secondary: #6b7280;
    --success: #10b981;
    --danger: #ef4444;
    --warning: #f59e0b;
    --info: #3b82f6;
    --light: #f9fafb;
    --dark: #1f2937;
    --card-bg: #ffffff;
    --border: #e5e7eb;
}

* {
    font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
}

body {
    background: linear-gradient(180deg, #f6f8fb 0%, #ffffff 100%);
    color: var(--dark);
}

.stApp {
    max-width: 1600px;
    margin: 0 auto;
}

.header {
    display: flex;
    align-items: center;
    gap: 16px;
    padding: 16px 20px;
    border-radius: 12px;
    margin-bottom: 16px;
    background: var(--card-bg);
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

.app-title {
    font-size: 28px;
    font-weight: 700;
    color: var(--primary);
    margin: 0;
}

.app-sub {
    color: var(--secondary);
    font-size: 14px;
    margin: 4px 0 0 0;
}

.kpi {
    background: var(--card-bg);
    border-radius: 12px;
    padding: 16px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
    text-align: center;
    height: 100%;
}

.kpi .value {
    font-size: 20px;
    font-weight: 700;
    margin-bottom: 4px;
}

.kpi .label {
    font-size: 12px;
    color: var(--secondary);
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.section {
    background: var(--card-bg);
    padding: 20px;
    border-radius: 12px;
    margin-bottom: 16px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

.section-title {
    font-size: 18px;
    font-weight: 600;
    color: var(--primary);
    margin-bottom: 16px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
}

.small {
    font-size: 12px;
    color: var(--secondary);
}

.code {
    background: #f3f4f6;
    padding: 8px 12px;
    border-radius: 6px;
    font-family: 'Monaco', 'Menlo', 'Ubuntu Mono', monospace;
    font-size: 13px;
    border: 1px solid var(--border);
}

.card {
    border-radius: 12px;
    padding: 16px;
    background: var(--card-bg);
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    margin-bottom: 16px;
    border: 1px solid var(--border);
}

.plotly-graph-div {
    border-radius: 12px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

/* Custom tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
}

.stTabs [data-baseweb="tab"] {
    background: #f1f5f9;
    border-radius: 8px 8px 0 0;
    padding: 10px 20px;
    font-weight: 500;
}

.stTabs [aria-selected="true"] {
    background: var(--primary);
    color: white;
}

/* Button styling */
.stButton button {
    border-radius: 8px;
    padding: 8px 16px;
    font-weight: 500;
}

/* Dataframe styling */
.dataframe {
    border-radius: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
}

/* Metric cards */
[data-testid="stMetric"] {
    background: var(--card-bg);
    border-radius: 12px;
    padding: 16px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

/* Sidebar */
.css-1d391kg {
    background: #f8fafc;
}

/* Custom colors for direction indicators */
.bullish {
    color: var(--success);
    font-weight: 600;
}

.bearish {
    color: var(--danger);
    font-weight: 600;
}

.neutral {
    color: var(--warning);
    font-weight: 600;
}
</style>
"""

# ---------------- UTILITY: NSE Session ----------------
@st.cache_data(ttl=300)
def get_nse_session():
    """
    Create a requests session with common NSE headers to fetch option chain.
    Using a session helps maintain cookies and reduces repeated costs.
    """
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": "https://www.nseindia.com/option-chain",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1"
    })
    # Try a warm-up GET to set cookies (may fail inside some sandboxed envs)
    try:
        s.get("https://www.nseindia.com/option-chain", timeout=8)
    except Exception:
        pass
    return s

# ---------------- FETCH / PARSE ----------------
def fetch_option_chain(symbol, session):
    """
    Fetch option-chain JSON for a symbol. Retries 3 times with backoff.
    Returns JSON or None.
    """
    url = (f"https://www.nseindia.com/api/option-chain-indices?symbol={symbol}"
           if symbol in INDICES else f"https://www.nseindia.com/api/option-chain-equities?symbol={symbol}")
    for attempt in range(3):
        try:
            resp = session.get(url, timeout=30)
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            # backoff with jitter
            time.sleep(random.uniform(1.0, 2.5) * (attempt + 1))
    return None

def parse_data(symbol, data):
    """
    Convert NSE option-chain JSON to a DataFrame for the nearest expiry.
    Columns: STRIKE, CALL_OI, CALL_CHNG_IN_OI, CALL_IV, CALL_LTP, PUT_OI, PUT_CHNG_IN_OI, PUT_IV, PUT_LTP
    """
    if not data:
        return pd.DataFrame()
    records = data.get("records", {}) or {}
    expiry_dates = records.get("expiryDates", []) or []
    if not expiry_dates:
        return pd.DataFrame()
    expiry = expiry_dates[0]
    rows = []
    for item in records.get("data", []) or []:
        if item.get("expiryDate") != expiry:
            continue
        ce = item.get("CE") or {}
        pe = item.get("PE") or {}
        rows.append({
            "STRIKE": item.get("strikePrice"),
            "CALL_OI": ce.get("openInterest", 0),
            "CALL_CHNG_IN_OI": ce.get("changeinOpenInterest", 0),
            "CALL_IV": ce.get("impliedVolatility", 0),
            "CALL_LTP": ce.get("lastPrice", 0),
            "PUT_OI": pe.get("openInterest", 0),
            "PUT_CHNG_IN_OI": pe.get("changeinOpenInterest", 0),
            "PUT_IV": pe.get("impliedVolatility", 0),
            "PUT_LTP": pe.get("lastPrice", 0)
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values("STRIKE").reset_index(drop=True)

# ---------------- ENHANCED ANALYTICS ----------------
def calculate_analytics(df, spot_price=None):
    """
    Enhanced analytics:
    - total/ATM PCR
    - true max pain (loss-based)
    - support/resistance by OI buildup
    - IV skew and approx expected 30-day move
    - directional score
    - volatility surface analysis
    - option flow analysis
    Returns a dict with rich metrics and modified df (with extras).
    """
    if df.empty:
        return {}
    df = df.copy()
    # estimate spot if not provided: strike with smallest difference between call & put OI
    if not spot_price or spot_price == 0:
        df['OI_DIFF'] = (df['CALL_OI'] - df['PUT_OI']).abs()
        spot_price = int(df.loc[df['OI_DIFF'].idxmin(), "STRIKE"])
    df['TOTAL_OI'] = df['CALL_OI'] + df['PUT_OI']
    df['OI_RATIO'] = df['PUT_OI'] / (df['CALL_OI'] + 1e-10)  # Avoid division by zero
    df['DELTA_CALL'] = df['CALL_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['DELTA_PUT'] = df['PUT_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['IV_DIFF'] = df['CALL_IV'] - df['PUT_IV']
    df['PRICE_RATIO'] = df['CALL_LTP'] / (df['PUT_LTP'] + 1e-10)
    
    total_call_oi = df['CALL_OI'].sum()
    total_put_oi = df['PUT_OI'].sum()
    pcr = total_put_oi / total_call_oi if total_call_oi > 0 else 0

    # ATM window: pick nearest strike and +/- 3 strikes (adjust for strike step)
    strike_step = int(df['STRIKE'].diff().median() or 50)
    atm_strike = int(df.iloc[(df['STRIKE'] - spot_price).abs().argsort()[:1]]['STRIKE'].values[0])
    window = 3 * strike_step
    window_df = df[(df['STRIKE'] >= atm_strike - window) & (df['STRIKE'] <= atm_strike + window)]
    atm_pcr = window_df['PUT_OI'].sum() / window_df['CALL_OI'].sum() if window_df['CALL_OI'].sum() > 0 else 0

    # true max pain computation (min aggregated option seller loss)
    strikes = df['STRIKE'].values
    losses = []
    for k in strikes:
        call_loss = ((np.clip(k - strikes, 0, None)) * df['CALL_OI']).sum()
        put_loss = ((np.clip(strikes - k, 0, None)) * df['PUT_OI']).sum()
        losses.append(call_loss + put_loss)
    max_pain = int(strikes[np.argmin(losses)]) if len(losses) > 0 else 0

    # support/resistance by OI buildup and current OI
    df['PUT_SCORE'] = df['PUT_OI'] * (1 + df['PUT_CHNG_IN_OI'] / (df['PUT_OI'].replace(0, 1) + 1))
    df['CALL_SCORE'] = df['CALL_OI'] * (1 + df['CALL_CHNG_IN_OI'] / (df['CALL_OI'].replace(0, 1) + 1))
    strongest_support = int(df.loc[df['PUT_SCORE'].idxmax(), 'STRIKE'])
    strongest_resistance = int(df.loc[df['CALL_SCORE'].idxmax(), 'STRIKE'])

    # IV metrics
    avg_call_iv = df['CALL_IV'].mean()
    avg_put_iv = df['PUT_IV'].mean()
    iv_skew = round(avg_call_iv - avg_put_iv, 2)
    iv_atm = window_df[['CALL_IV', 'PUT_IV']].mean().mean() if not window_df.empty else (avg_call_iv + avg_put_iv) / 2
    expected_move_annual = iv_atm / 100.0
    # approximate 30d expected move (sqrt scaling for monthly ~ sqrt(1/12) of annual vol)
    expected_move_30d = round(spot_price * expected_move_annual / np.sqrt(12), 2)
    
    # Volatility surface analysis
    iv_slope_call = np.polyfit(df['STRIKE'], df['CALL_IV'], 1)[0] if len(df) > 1 else 0
    iv_slope_put = np.polyfit(df['STRIKE'], df['PUT_IV'], 1)[0] if len(df) > 1 else 0
    
    # Option flow analysis
    call_flow = (df['CALL_CHNG_IN_OI'] * df['CALL_LTP']).sum()
    put_flow = (df['PUT_CHNG_IN_OI'] * df['PUT_LTP']).sum()
    net_flow = call_flow - put_flow
    flow_ratio = abs(call_flow / put_flow) if put_flow != 0 else float('inf')

    # volume ratio and direction score
    call_vol = df['CALL_CHNG_IN_OI'].sum()
    put_vol = df['PUT_CHNG_IN_OI'].sum()
    vol_ratio = (call_vol + 1) / (put_vol + 1)
    dir_score = 0.0
    dir_score += (pcr - 1.0) * 2.0
    dir_score += (iv_skew / 5.0)
    dir_score += (vol_ratio - 1.0)
    dir_score += (net_flow / max(abs(net_flow), 1)) * 0.5  # Add flow impact
    
    if dir_score > 1.2:
        direction = "Bullish"
    elif dir_score < -1.2:
        direction = "Bearish"
    else:
        direction = "Neutral"

    top_calls = df.nlargest(5, 'CALL_OI')[['STRIKE', 'CALL_OI', 'CALL_IV', 'CALL_CHNG_IN_OI', 'CALL_LTP']]
    top_puts = df.nlargest(5, 'PUT_OI')[['STRIKE', 'PUT_OI', 'PUT_IV', 'PUT_CHNG_IN_OI', 'PUT_LTP']]
    
    # Option clustering for pattern recognition
    if len(df) > 10:
        try:
            # Prepare data for clustering
            cluster_data = df[['STRIKE', 'CALL_OI', 'PUT_OI', 'CALL_IV', 'PUT_IV']].copy()
            cluster_data = (cluster_data - cluster_data.mean()) / cluster_data.std()
            
            # Apply K-means clustering
            kmeans = KMeans(n_clusters=3, random_state=42, n_init=10)
            df['CLUSTER'] = kmeans.fit_predict(cluster_data)
        except:
            df['CLUSTER'] = 0

    return {
        "df": df,
        "spot_price": spot_price,
        "pcr": round(pcr, 2),
        "pcr_atm": round(atm_pcr, 2),
        "max_pain": max_pain,
        "support": strongest_support,
        "resistance": strongest_resistance,
        "iv_skew": iv_skew,
        "iv_slope_call": round(iv_slope_call, 4),
        "iv_slope_put": round(iv_slope_put, 4),
        "expected_move_30d": expected_move_30d,
        "direction": direction,
        "dir_score": round(dir_score, 2),
        "call_flow": round(call_flow, 2),
        "put_flow": round(put_flow, 2),
        "net_flow": round(net_flow, 2),
        "flow_ratio": round(flow_ratio, 2),
        "top_calls": top_calls,
        "top_puts": top_puts
    }

# ---------------- ML: Regression & Classification ----------------
def train_ml_models_regression(df):
    """
    Regression models to predict 'attractive' strikes (explainable).
    Returns models' results dict and recommended top calls/puts by prediction.
    """
    if df.empty or len(df) < 10:
        return {}, [], []
    
    # Create target variable based on OI and price action
    df_ml = df.copy()
    df_ml['TARGET'] = df_ml['CALL_OI'] * df_ml['CALL_LTP'] - df_ml['PUT_OI'] * df_ml['PUT_LTP']
    
    X = df_ml[['CALL_OI', 'PUT_OI', 'CALL_CHNG_IN_OI', 'PUT_CHNG_IN_OI', 'CALL_IV', 'PUT_IV', 'CALL_LTP', 'PUT_LTP']].fillna(0)
    y = df_ml['TARGET']
    
    # Feature importance using Random Forest
    rf = RandomForestRegressor(n_estimators=100, random_state=42)
    rf.fit(X, y)
    feature_importance = dict(zip(X.columns, rf.feature_importances_))
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    scaler = StandardScaler()
    X_train_s = scaler.fit_transform(X_train)
    X_test_s = scaler.transform(X_test)

    models = {
        "Linear": LinearRegression(),
        "Ridge": Ridge(alpha=1.0),
        "Lasso": Lasso(alpha=0.01),
        "RF": RandomForestRegressor(n_estimators=150, random_state=42),
        "GB": GradientBoostingRegressor(n_estimators=150, random_state=42)
    }
    results = {}
    for name, model in models.items():
        model.fit(X_train_s, y_train)
        y_pred = model.predict(X_test_s)
        mae = mean_absolute_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)
        results[name] = {"model": model, "mae": mae, "r2": r2, "scaler": scaler}

    best_name = min(results, key=lambda k: results[k]['mae'])
    best_model = results[best_name]['model']
    best_scaler = results[best_name]['scaler']

    X_full = scaler.transform(X)
    df_ml['ML_PREDICTED_VALUE'] = best_model.predict(X_full)
    
    # Get top calls and puts based on ML predictions
    top_calls = df_ml.nlargest(3, 'ML_PREDICTED_VALUE')['STRIKE'].tolist()
    top_puts = df_ml.nsmallest(3, 'ML_PREDICTED_VALUE')['STRIKE'].tolist()

    return results, top_calls, top_puts, feature_importance

def train_ml_models_classification(df):
    """
    Classification to detect bias from options features.
    Simple labeled by PCR thresholds.
    """
    if df.empty or len(df) < 12:
        return {"RF": 0.0, "LR": 0.0}, "Neutral", {}
    
    d = df.copy()
    d['PCR'] = d['PUT_OI'] / (d['CALL_OI'] + 1)
    d['IVS'] = d['CALL_IV'] - d['PUT_IV']
    d['FLOW_RATIO'] = (d['CALL_CHNG_IN_OI'] * d['CALL_LTP']) / (d['PUT_CHNG_IN_OI'] * d['PUT_LTP'] + 1)
    
    features = ['CALL_OI', 'PUT_OI', 'PCR', 'IVS', 'CALL_CHNG_IN_OI', 'PUT_CHNG_IN_OI', 'FLOW_RATIO']
    d['LABEL'] = np.where(d['PCR'] > 1.2, 'Bullish', np.where(d['PCR'] < 0.8, 'Bearish', 'Neutral'))
    
    X = d[features].fillna(0)
    y = d['LABEL']
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
    
    rf = RandomForestClassifier(n_estimators=150, random_state=42)
    rf.fit(X_train, y_train)
    acc_rf = accuracy_score(y_test, rf.predict(X_test))
    feature_importance = dict(zip(features, rf.feature_importances_))
    
    lr = LogisticRegression(max_iter=500, class_weight='balanced')
    lr.fit(X_train, y_train)
    acc_lr = accuracy_score(y_test, lr.predict(X_test))
    
    # consensus label from RF predictions on the set (mode)
    preds = rf.predict(X)
    consensus = pd.Series(preds).mode().iloc[0] if len(preds) > 0 else "Neutral"
    
    # Confusion matrix and classification report
    y_pred = rf.predict(X_test)
    cm = confusion_matrix(y_test, y_pred, labels=['Bullish', 'Neutral', 'Bearish'])
    cr = classification_report(y_test, y_pred, output_dict=True)
    
    return {"RF": round(acc_rf, 3), "LR": round(acc_lr, 3)}, consensus, {
        "feature_importance": feature_importance,
        "confusion_matrix": cm,
        "classification_report": cr
    }

# ---------------- ADVANCED CHARTS ----------------
def create_oi_chart(df, line_shape='spline'):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['CALL_OI'], mode='lines', name='Call OI', 
                             line=dict(shape=line_shape, width=3, color='#FF6B6B')))
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['PUT_OI'], mode='lines', name='Put OI', 
                             line=dict(shape=line_shape, width=3, color='#4ECDC4')))
    fig.update_layout(
        title="Open Interest Distribution",
        xaxis_title="Strike Price",
        yaxis_title="Open Interest",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_sentiment_chart(df, line_shape='spline'):
    df_local = df.copy()
    df_local['SENT'] = df_local['CALL_OI'] - df_local['PUT_OI']
    fig = go.Figure([go.Bar(x=df_local['STRIKE'], y=df_local['SENT'], 
                           marker_color=np.where(df_local['SENT'] > 0, '#FF6B6B', '#4ECDC4'))])
    fig.update_layout(
        title="Sentiment (Call OI - Put OI)",
        xaxis_title="Strike Price",
        yaxis_title="Call-Put OI Difference",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_iv_comparison_chart(df, line_shape='spline'):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['CALL_IV'], name='Call IV', 
                             line=dict(shape=line_shape, width=3, color='#FF6B6B')))
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['PUT_IV'], name='Put IV', 
                             line=dict(shape=line_shape, width=3, color='#4ECDC4')))
    fig.update_layout(
        title="Implied Volatility (Call vs Put)",
        xaxis_title="Strike Price",
        yaxis_title="Implied Volatility (%)",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_volatility_surface_chart(df):
    fig = make_subplots(rows=1, cols=2, subplot_titles=("Call IV Surface", "Put IV Surface"))
    
    # Call IV surface
    fig.add_trace(
        go.Scatter(x=df['STRIKE'], y=df['CALL_IV'], mode='lines+markers', 
                  name='Call IV', line=dict(width=3, color='#FF6B6B')),
        row=1, col=1
    )
    
    # Put IV surface
    fig.add_trace(
        go.Scatter(x=df['STRIKE'], y=df['PUT_IV'], mode='lines+markers', 
                  name='Put IV', line=dict(width=3, color='#4ECDC4')),
        row=1, col=2
    )
    
    fig.update_xaxes(title_text="Strike Price", row=1, col=1)
    fig.update_xaxes(title_text="Strike Price", row=1, col=2)
    fig.update_yaxes(title_text="Implied Volatility (%)", row=1, col=1)
    fig.update_yaxes(title_text="Implied Volatility (%)", row=1, col=2)
    
    fig.update_layout(
        title="Volatility Surface Analysis",
        height=400,
        margin=dict(t=80, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    
    return fig

def create_ml_prediction_chart(df, analytics, top_calls, top_puts, line_shape='spline'):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['TOTAL_OI'], mode='lines+markers', 
                             name='Total OI', line=dict(shape=line_shape, width=3, color='#45B7D1')))
    
    if 'ML_PREDICTED_VALUE' in df.columns:
        fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['ML_PREDICTED_VALUE'], mode='markers', 
                                name='ML Predicted', marker=dict(size=8, color='#FBE555')))
    
    fig.add_vline(x=analytics.get('max_pain', None), line_dash='dash', line_color='#964B00', 
                  annotation_text='Max Pain', annotation_position="top right")
    
    for s in top_calls:
        fig.add_vline(x=s, line_dash='dot', line_color='green', annotation_text=f'Call {s}')
    
    for s in top_puts:
        fig.add_vline(x=s, line_dash='dot', line_color='red', annotation_text=f'Put {s}')
    
    fig.update_layout(
        title="ML Predicted Strikes & OI",
        xaxis_title="Strike Price",
        yaxis_title="Total OI / ML Prediction",
        height=450,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_model_performance_chart(ml_results):
    if not ml_results:
        return go.Figure()
    
    models = list(ml_results.keys())
    mae = [ml_results[m]['mae'] for m in models]
    r2 = [ml_results[m]['r2'] for m in models]
    
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    
    # Bar for MAE
    fig.add_trace(
        go.Bar(
            x=models,
            y=mae,
            name='MAE',
            marker_color='orange'
        ),
        secondary_y=False
    )
    
    # Line for R2
    fig.add_trace(
        go.Scatter(
            x=models,
            y=r2,
            name='R²',
            mode='lines+markers',
            line=dict(color='blue', width=2),
            marker=dict(size=8)
        ),
        secondary_y=True
    )
    
    # Layout with dual y-axis
    fig.update_layout(
        title="Regression Model Performance",
        xaxis_title="Model",
        yaxis=dict(title="MAE", side='left', showgrid=False),
        yaxis2=dict(title="R²", side='right', showgrid=False, range=[-1, 1]),
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    
    return fig

def create_feature_importance_chart(feature_importance, title="Feature Importance"):
    features = list(feature_importance.keys())
    importance = list(feature_importance.values())
    
    # Sort by importance
    sorted_idx = np.argsort(importance)
    features = [features[i] for i in sorted_idx]
    importance = [importance[i] for i in sorted_idx]
    
    fig = go.Figure(go.Bar(
        x=importance,
        y=features,
        orientation='h',
        marker_color='#45B7D1'
    ))
    
    fig.update_layout(
        title=title,
        xaxis_title="Importance",
        yaxis_title="Features",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    
    return fig

def create_option_flow_chart(analytics):
    flow_data = {
        'Type': ['Call Flow', 'Put Flow', 'Net Flow'],
        'Value': [analytics.get('call_flow', 0), analytics.get('put_flow', 0), analytics.get('net_flow', 0)]
    }
    
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1']
    
    fig = go.Figure(go.Bar(
        x=flow_data['Type'],
        y=flow_data['Value'],
        marker_color=colors,
        text=flow_data['Value'],
        textposition='auto'
    ))
    
    fig.update_layout(
        title="Option Flow Analysis",
        xaxis_title="Flow Type",
        yaxis_title="Value (₹)",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    
    return fig

def create_cluster_analysis_chart(df):
    if 'CLUSTER' not in df.columns:
        return go.Figure()
    
    fig = go.Figure()
    
    for cluster in sorted(df['CLUSTER'].unique()):
        cluster_data = df[df['CLUSTER'] == cluster]
        fig.add_trace(go.Scatter(
            x=cluster_data['STRIKE'],
            y=cluster_data['TOTAL_OI'],
            mode='markers',
            name=f'Cluster {cluster}',
            marker=dict(size=8)
        ))
    
    fig.update_layout(
        title="Option Chain Clustering Analysis",
        xaxis_title="Strike Price",
        yaxis_title="Total OI",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    
    return fig

# ---------------- EXPORT TO EXCEL ----------------
def export_to_excel(df, analytics, symbol, ml_results=None, top_calls=None, top_puts=None):
    """
    Export option chain data and analytics to Excel with professional formatting.
    """
    try:
        # Create a workbook and get the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{symbol} Option Chain"
        
        # Add title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"{symbol} Option Chain Analysis - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        title_cell.font = Font(bold=True, size=16, color="0047AB")
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add analytics section
        ws['A3'] = "ANALYTICS SUMMARY"
        ws['A3'].font = Font(bold=True, size=14, color="333333")
        
        analytics_data = [
            ["Spot Price", analytics.get('spot_price', 'N/A')],
            ["PCR", analytics.get('pcr', 'N/A')],
            ["ATM PCR", analytics.get('pcr_atm', 'N/A')],
            ["Max Pain", analytics.get('max_pain', 'N/A')],
            ["Support", analytics.get('support', 'N/A')],
            ["Resistance", analytics.get('resistance', 'N/A')],
            ["IV Skew", analytics.get('iv_skew', 'N/A')],
            ["Expected 30D Move", analytics.get('expected_move_30d', 'N/A')],
            ["Direction", analytics.get('direction', 'N/A')],
            ["Direction Score", analytics.get('dir_score', 'N/A')],
            ["Call Flow", analytics.get('call_flow', 'N/A')],
            ["Put Flow", analytics.get('put_flow', 'N/A')],
            ["Net Flow", analytics.get('net_flow', 'N/A')],
            ["Flow Ratio", analytics.get('flow_ratio', 'N/A')]
        ]
        
        for i, (label, value) in enumerate(analytics_data, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Add ML predictions if available
        if ml_results and top_calls and top_puts:
            ws['A18'] = "ML PREDICTIONS"
            ws['A18'].font = Font(bold=True, size=14, color="333333")
            
            ws['A19'] = "Top Call Strikes"
            ws['A19'].font = Font(bold=True)
            for i, strike in enumerate(top_calls, start=19):
                ws[f'B{i}'] = strike
            
            ws['A23'] = "Top Put Strikes"
            ws['A23'].font = Font(bold=True)
            for i, strike in enumerate(top_puts, start=23):
                ws[f'B{i}'] = strike
        
        # Add option chain data
        start_row = 27
        ws[f'A{start_row}'] = "OPTION CHAIN DATA"
        ws[f'A{start_row}'].font = Font(bold=True, size=14, color="333333")
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row+1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0047AB", end_color="0047AB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row+2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Format based on values
                if headers[col_idx-1] == 'CALL_CHNG_IN_OI' and value > 0:
                    cell.font = Font(color="008000")  # Green for positive change
                elif headers[col_idx-1] == 'CALL_CHNG_IN_OI' and value < 0:
                    cell.font = Font(color="FF0000")  # Red for negative change
                
                if headers[col_idx-1] == 'PUT_CHNG_IN_OI' and value > 0:
                    cell.font = Font(color="008000")  # Green for positive change
                elif headers[col_idx-1] == 'PUT_CHNG_IN_OI' and value < 0:
                    cell.font = Font(color="FF0000")  # Red for negative change
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        filename = f"{symbol}_Option_Chain_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(SAVE_FOLDER, filename)
        wb.save(filepath)
        
        return filepath, None
        
    except Exception as e:
        return None, f"Excel export failed: {str(e)}"

# ---------------- STREAMLIT UI ----------------
def main():
    # Inject custom CSS
    st.markdown(PRO_CSS, unsafe_allow_html=True)
    
    # Header
    col1, col2 = st.columns([1, 4])
    with col1:
        st.image("https://www.nseindia.com/assets/images/nse-logo.svg", width=80)
    with col2:
        st.markdown('<h1 class="app-title">NSE Option Chain Analyzer</h1>', unsafe_allow_html=True)
        st.markdown('<p class="app-sub">Professional-grade analytics, ML predictions, and strategy insights</p>', unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown("## Settings")
        symbol = st.text_input("Symbol (e.g., NIFTY, BANKNIFTY, RELIANCE, INFY)", "NIFTY").upper()
        refresh_interval = st.slider("Refresh Interval (seconds)", 10, 300, 60)
        auto_refresh = st.checkbox("Auto Refresh", value=True)
        
        st.markdown("## Analysis Options")
        show_ml = st.checkbox("Show ML Predictions", value=True)
        show_clusters = st.checkbox("Show Clustering", value=True)
        show_flow = st.checkbox("Show Option Flow", value=True)
        
        st.markdown("## Export Options")
        export_excel = st.button("Export to Excel")
        
        st.markdown("---")
        st.markdown("### About")
        st.markdown("""
        This app provides professional-grade analysis of NSE option chains with:
        - Real-time data fetching
        - Advanced analytics (PCR, Max Pain, Support/Resistance)
        - Machine learning predictions
        - Volatility analysis
        - Option flow tracking
        - Professional Excel exports
        """)
    
    # Initialize session state
    if 'last_refresh' not in st.session_state:
        st.session_state.last_refresh = 0
    if 'session' not in st.session_state:
        st.session_state.session = get_nse_session()
    if 'data' not in st.session_state:
        st.session_state.data = {}
    if 'analytics' not in st.session_state:
        st.session_state.analytics = {}
    if 'ml_results' not in st.session_state:
        st.session_state.ml_results = {}
    if 'top_calls' not in st.session_state:
        st.session_state.top_calls = []
    if 'top_puts' not in st.session_state:
        st.session_state.top_puts = []
    if 'classification_results' not in st.session_state:
        st.session_state.classification_results = {}
    
    # Refresh logic
    current_time = time.time()
    should_refresh = (current_time - st.session_state.last_refresh > refresh_interval and auto_refresh) or st.button("Refresh Now")
    
    if should_refresh:
        with st.spinner(f"Fetching {symbol} option chain..."):
            data = fetch_option_chain(symbol, st.session_state.session)
            if data:
                df = parse_data(symbol, data)
                if not df.empty:
                    st.session_state.data[symbol] = df
                    analytics = calculate_analytics(df)
                    st.session_state.analytics[symbol] = analytics
                    
                    # ML predictions
                    if show_ml:
                        ml_results, top_calls, top_puts, feature_importance = train_ml_models_regression(analytics['df'])
                        st.session_state.ml_results[symbol] = ml_results
                        st.session_state.top_calls = top_calls
                        st.session_state.top_puts = top_puts
                        st.session_state.feature_importance = feature_importance
                        
                        # Classification
                        classification_results, consensus, details = train_ml_models_classification(analytics['df'])
                        st.session_state.classification_results[symbol] = classification_results
                        st.session_state.consensus = consensus
                
                st.session_state.last_refresh = current_time
                st.success(f"Data refreshed for {symbol}")
            else:
                st.error(f"Failed to fetch data for {symbol}")
    
    # Display data if available
    if symbol in st.session_state.data and not st.session_state.data[symbol].empty:
        df = st.session_state.data[symbol]
        analytics = st.session_state.analytics.get(symbol, {})
        
        # KPI Cards
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.markdown('<div class="kpi">', unsafe_allow_html=True)
            st.metric("Spot Price", f"₹{analytics.get('spot_price', 'N/A')}")
            st.markdown('</div>', unsafe_allow_html=True)
        with col2:
            st.markdown('<div class="kpi">', unsafe_allow_html=True)
            st.metric("PCR", analytics.get('pcr', 'N/A'))
            st.markdown('</div>', unsafe_allow_html=True)
        with col3:
            st.markdown('<div class="kpi">', unsafe_allow_html=True)
            st.metric("Max Pain", f"₹{analytics.get('max_pain', 'N/A')}")
            st.markdown('</div>', unsafe_allow_html=True)
        with col4:
            st.markdown('<div class="kpi">', unsafe_allow_html=True)
            direction = analytics.get('direction', 'Neutral')
            color_class = "bullish" if direction == "Bullish" else "bearish" if direction == "Bearish" else "neutral"
            st.markdown(f'<div class="value {color_class}">{direction}</div>', unsafe_allow_html=True)
            st.markdown('<div class="label">Direction</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with col5:
            st.markdown('<div class="kpi">', unsafe_allow_html=True)
            st.metric("30D Expected Move", f"₹{analytics.get('expected_move_30d', 'N/A')}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Tabs for different sections
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["Overview", "Analytics", "ML Predictions", "Option Chain", "Export"])
        
        with tab1:
            st.markdown('<div class="section">', unsafe_allow_html=True)
            st.markdown('<h2 class="section-title">Overview</h2>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.plotly_chart(create_oi_chart(df), use_container_width=True)
            with col2:
                st.plotly_chart(create_sentiment_chart(df), use_container_width=True)
            
            col3, col4 = st.columns(2)
            with col3:
                st.plotly_chart(create_iv_comparison_chart(df), use_container_width=True)
            with col4:
                st.plotly_chart(create_volatility_surface_chart(df), use_container_width=True)
            
            if show_flow:
                st.plotly_chart(create_option_flow_chart(analytics), use_container_width=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        with tab2:
            st.markdown('<div class="section">', unsafe_allow_html=True)
            st.markdown('<h2 class="section-title">Advanced Analytics</h2>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("### Support & Resistance")
                st.markdown(f"**Strongest Support:** ₹{analytics.get('support', 'N/A')}")
                st.markdown(f"**Strongest Resistance:** ₹{analytics.get('resistance', 'N/A')}")
                
                st.markdown("### Volatility Analysis")
                st.markdown(f"**IV Skew (Call - Put):** {analytics.get('iv_skew', 'N/A')}")
                st.markdown(f"**Call IV Slope:** {analytics.get('iv_slope_call', 'N/A')}")
                st.markdown(f"**Put IV Slope:** {analytics.get('iv_slope_put', 'N/A')}")
                
                st.markdown("### Option Flow")
                st.markdown(f"**Call Flow:** ₹{analytics.get('call_flow', 'N/A'):,}")
                st.markdown(f"**Put Flow:** ₹{analytics.get('put_flow', 'N/A'):,}")
                st.markdown(f"**Net Flow:** ₹{analytics.get('net_flow', 'N/A'):,}")
                st.markdown(f"**Flow Ratio:** {analytics.get('flow_ratio', 'N/A')}")
            
            with col2:
                st.markdown("### Top Calls by OI")
                st.dataframe(analytics.get('top_calls', pd.DataFrame()), use_container_width=True)
                
                st.markdown("### Top Puts by OI")
                st.dataframe(analytics.get('top_puts', pd.DataFrame()), use_container_width=True)
            
            if show_clusters and 'CLUSTER' in analytics.get('df', pd.DataFrame()).columns:
                st.plotly_chart(create_cluster_analysis_chart(analytics['df']), use_container_width=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        with tab3:
            if show_ml and symbol in st.session_state.ml_results:
                st.markdown('<div class="section">', unsafe_allow_html=True)
                st.markdown('<h2 class="section-title">Machine Learning Predictions</h2>', unsafe_allow_html=True)
                
                ml_results = st.session_state.ml_results[symbol]
                classification_results = st.session_state.classification_results.get(symbol, {})
                
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("### Regression Performance")
                    st.plotly_chart(create_model_performance_chart(ml_results), use_container_width=True)
                
                with col2:
                    st.markdown("### Feature Importance")
                    st.plotly_chart(create_feature_importance_chart(
                        st.session_state.feature_importance, 
                        "Regression Feature Importance"
                    ), use_container_width=True)
                
                st.markdown("### ML Recommended Strikes")
                col3, col4 = st.columns(2)
                with col3:
                    st.markdown("#### Top Call Strikes")
                    for strike in st.session_state.top_calls:
                        st.markdown(f"- ₹{strike}")
                with col4:
                    st.markdown("#### Top Put Strikes")
                    for strike in st.session_state.top_puts:
                        st.markdown(f"- ₹{strike}")
                
                st.plotly_chart(create_ml_prediction_chart(
                    analytics['df'], analytics, 
                    st.session_state.top_calls, st.session_state.top_puts
                ), use_container_width=True)
                
                st.markdown("### Classification Results")
                col5, col6 = st.columns(2)
                with col5:
                    st.markdown(f"**Random Forest Accuracy:** {classification_results.get('RF', 0)}")
                    st.markdown(f"**Logistic Regression Accuracy:** {classification_results.get('LR', 0)}")
                    st.markdown(f"**Consensus Bias:** {st.session_state.consensus}")
                
                with col6:
                    if 'feature_importance' in st.session_state.get('classification_details', {}):
                        st.plotly_chart(create_feature_importance_chart(
                            st.session_state.classification_details['feature_importance'],
                            "Classification Feature Importance"
                        ), use_container_width=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.info("Enable ML predictions in sidebar settings to see this section")
        
        with tab4:
            st.markdown('<div class="section">', unsafe_allow_html=True)
            st.markdown('<h2 class="section-title">Option Chain Data</h2>', unsafe_allow_html=True)
            
            # Format the dataframe for better display
            display_df = df.copy()
            display_df = display_df.round(2)
            
            # Add color formatting for changes in OI
            def color_positive_green(val):
                color = 'green' if val > 0 else 'red' if val < 0 else 'black'
                return f'color: {color}'
            
            styled_df = display_df.style.applymap(color_positive_green, subset=['CALL_CHNG_IN_OI', 'PUT_CHNG_IN_OI'])
            st.dataframe(styled_df, use_container_width=True, height=600)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        with tab5:
            st.markdown('<div class="section">', unsafe_allow_html=True)
            st.markdown('<h2 class="section-title">Export Data</h2>', unsafe_allow_html=True)
            
            if export_excel:
                with st.spinner("Exporting to Excel..."):
                    filepath, error = export_to_excel(
                        df, analytics, symbol, 
                        st.session_state.ml_results.get(symbol), 
                        st.session_state.top_calls, 
                        st.session_state.top_puts
                    )
                    
                    if filepath:
                        st.success(f"Excel file saved to: {filepath}")
                        
                        # Provide download link
                        with open(filepath, "rb") as f:
                            bytes_data = f.read()
                        st.download_button(
                            label="Download Excel File",
                            data=bytes_data,
                            file_name=os.path.basename(filepath),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.error(f"Export failed: {error}")
            
            st.markdown("""
            ### Export Options
            
            Click the button above to export the current option chain data and analytics to Excel.
            
            The exported file will include:
            - Analytics summary (PCR, Max Pain, Support/Resistance, etc.)
            - ML predictions (if enabled)
            - Complete option chain data
            - Professional formatting
            """)
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    else:
        st.info("Enter a symbol and click 'Refresh Now' to load data")
    
    # Auto-refresh countdown
    if auto_refresh:
        next_refresh = max(0, refresh_interval - (time.time() - st.session_state.last_refresh))
        st.sidebar.markdown(f"Next refresh in: {int(next_refresh)} seconds")
        
        # This will trigger a rerun when the countdown reaches 0
        if next_refresh <= 0:
            st.rerun()
        else:
            time.sleep(0.5)
            st.rerun()

if __name__ == "__main__":
    main()