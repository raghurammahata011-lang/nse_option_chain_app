# utils/excel_export.py
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
import plotly.io as pio

def create_excel_export(df, analytics, symbol, ml_results, top_calls, top_puts, session):
    """
    Create a professional Excel workbook with multiple sheets, formatting, and charts.
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Data Sheet
    ws_data = wb.create_sheet("Option Chain Data")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    # Format data sheet
    for cell in ws_data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=1, max_col=ws_data.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Analytics Sheet
    ws_analytics = wb.create_sheet("Analytics")
    analytics_data = [
        ["Metric", "Value"],
        ["Symbol", symbol],
        ["Spot Price", analytics.get("spot_price", "N/A")],
        ["PCR (Total)", analytics.get("pcr", "N/A")],
        ["PCR (ATM)", analytics.get("pcr_atm", "N/A")],
        ["Max Pain", analytics.get("max_pain", "N/A")],
        ["Support", analytics.get("support", "N/A")],
        ["Resistance", analytics.get("resistance", "N/A")],
        ["IV Skew", analytics.get("iv_skew", "N/A")],
        ["IV Slope (Call)", analytics.get("iv_slope_call", "N/A")],
        ["IV Slope (Put)", analytics.get("iv_slope_put", "N/A")],
        ["Expected 30D Move", analytics.get("expected_move_30d", "N/A")],
        ["Direction", analytics.get("direction", "N/A")],
        ["Direction Score", analytics.get("dir_score", "N/A")],
        ["Call Flow", analytics.get("call_flow", "N/A")],
        ["Put Flow", analytics.get("put_flow", "N/A")],
        ["Net Flow", analytics.get("net_flow", "N/A")],
        ["Flow Ratio", analytics.get("flow_ratio", "N/A")],
        ["VIX-like Indicator", analytics.get("vix_like", "N/A")],
        ["Put-Call Parity Dev", analytics.get("put_call_parity_dev", "N/A")],
        ["Gamma Exposure", analytics.get("gamma_exposure", "N/A")]
    ]
    
    for row in analytics_data:
        ws_analytics.append(row)
    
    # Format analytics sheet
    for cell in ws_analytics[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row in ws_analytics.iter_rows(min_row=2, max_row=ws_analytics.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")
    
    # ML Results Sheet
    if ml_results:
        ws_ml = wb.create_sheet("ML Results")
        ml_data = [["Model", "MAE", "R²"]]
        for model_name, result in ml_results.items():
            ml_data.append([model_name, result.get("mae", "N/A"), result.get("r2", "N/A")])
        
        ml_data.append([])
        ml_data.append(["Top Calls (ML)", "Top Puts (ML)"])
        for i in range(max(len(top_calls), len(top_puts))):
            call_val = top_calls[i] if i < len(top_calls) else ""
            put_val = top_puts[i] if i < len(top_puts) else ""
            ml_data.append([call_val, put_val])
        
        for row in ml_data:
            ws_ml.append(row)
        
        # Format ML sheet
        for cell in ws_ml[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row in ws_ml.iter_rows(min_row=2, max_row=ws_ml.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    
    # Top OI Sheet
    ws_top = wb.create_sheet("Top OI")
    top_calls_data = analytics.get("top_calls", pd.DataFrame())
    top_puts_data = analytics.get("top_puts", pd.DataFrame())
    
    if not top_calls_data.empty:
        ws_top.append(["Top Calls by OI"])
        for r in dataframe_to_rows(top_calls_data, index=False, header=True):
            ws_top.append(r)
        
        ws_top.append([])
    
    if not top_puts_data.empty:
        ws_top.append(["Top Puts by OI"])
        for r in dataframe_to_rows(top_puts_data, index=False, header=True):
            ws_top.append(r)
    
    # Format Top OI sheet
    for row in ws_top.iter_rows():
        for cell in row:
            if cell.value in ["Top Calls by OI", "Top Puts by OI"]:
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer