# nse_option_chain_app_final.py
# Professional, full-featured Streamlit app for NSE option chain analytics,
# watchlist, strategy signals, and backtesting (underlying via yfinance).
#
# Requirements:
# streamlit, pandas, numpy, plotly, scikit-learn, openpyxl, requests, kaleido, yfinance

import time
import random
import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.io as pio
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from sklearn.linear_model import LinearRegression, Ridge, Lasso, LogisticRegression
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score, accuracy_score, confusion_matrix, classification_report
from sklearn.cluster import KMeans
from strategies import option_greeks, plot_strategy, strategy_summary
import os
import concurrent.futures
import yfinance as yf
import warnings
warnings.filterwarnings('ignore')

# ---------------- CONFIG ----------------
SAVE_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "NSE_STOCK")
os.makedirs(SAVE_FOLDER, exist_ok=True)
INDICES = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"]
# tune this if necessary
THREAD_WORKERS = 6

# ---------------- PROFESSIONAL CSS ----------------
PRO_CSS = """
<style>
:root {
    --primary: #0b69ff;
    --secondary: #6b7280;
    --success: #10b981;
    --danger: #ef4444;
    --warning: #f59e0b;
    --info: #3b82f6;
    --light: #f9fafb;
    --dark: #1f2937;
    --card-bg: #ffffff;
    --border: #e5e7eb;
}

* {
    font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
}

body {
    background: linear-gradient(180deg, #f6f8fb 0%, #ffffff 100%);
    color: var(--dark);
}

.stApp {
    max-width: 1800px;
    margin: 0 auto;
}

.header {
    display: flex;
    align-items: center;
    gap: 16px;
    padding: 16px 20px;
    border-radius: 12px;
    margin-bottom: 16px;
    background: var(--card-bg);
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

.app-title {
    font-size: 28px;
    font-weight: 700;
    color: var(--primary);
    margin: 0;
}

.app-sub {
    color: var(--secondary);
    font-size: 14px;
    margin: 4px 0 0 0;
}

.kpi {
    background: var(--card-bg);
    border-radius: 12px;
    padding: 16px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
    text-align: center;
    height: 100%;
}

.kpi .value {
    font-size: 20px;
    font-weight: 700;
    margin-bottom: 4px;
}

.kpi .label {
    font-size: 12px;
    color: var(--secondary);
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.section {
    background: var(--card-bg);
    padding: 20px;
    border-radius: 12px;
    margin-bottom: 16px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

.section-title {
    font-size: 18px;
    font-weight: 600;
    color: var(--primary);
    margin-bottom: 16px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
}

.small {
    font-size: 12px;
    color: var(--secondary);
}

.code {
    background: #f3f4f6;
    padding: 8px 12px;
    border-radius: 6px;
    font-family: 'Monaco', 'Menlo', 'Ubuntu Mono', monospace;
    font-size: 13px;
    border: 1px solid var(--border);
}

.card {
    border-radius: 12px;
    padding: 16px;
    background: var(--card-bg);
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    margin-bottom: 16px;
    border: 1px solid var(--border);
}

.plotly-graph-div {
    border-radius: 12px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

/* Custom tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
}

.stTabs [data-baseweb="tab"] {
    background: #f1f5f9;
    border-radius: 8px 8px 0 0;
    padding: 10px 20px;
    font-weight: 500;
}

.stTabs [aria-selected="true"] {
    background: var(--primary);
    color: white;
}

/* Button styling */
.stButton button {
    border-radius: 8px;
    padding: 8px 16px;
    font-weight: 500;
}

/* Dataframe styling */
.dataframe {
    border-radius: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
}

/* Metric cards */
[data-testid="stMetric"] {
    background: var(--card-bg);
    border-radius: 12px;
    padding: 16px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    border: 1px solid var(--border);
}

/* Sidebar */
.css-1d391kg {
    background: #f8fafc;
}

/* Custom colors for direction indicators */
.bullish {
    color: var(--success);
    font-weight: 600;
}

.bearish {
    color: var(--danger);
    font-weight: 600;
}

.neutral {
    color: var(--warning);
    font-weight: 600;
}

/* Custom scrollbar for tables */
.dataframe::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}

.dataframe::-webkit-scrollbar-track {
    background: #f1f1f1;
    border-radius: 10px;
}

.dataframe::-webkit-scrollbar-thumb {
    background: #c1c1c1;
    border-radius: 10px;
}

.dataframe::-webkit-scrollbar-thumb:hover {
    background: #a8a8a8;
}

/* Improved visibility for tables */
.dataframe th {
    background-color: #f0f7ff !important;
    position: sticky;
    top: 0;
    z-index: 10;
}

/* Enhanced visibility for important metrics */
.highlight-metric {
    background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
    border-left: 4px solid var(--primary);
    padding: 12px;
    border-radius: 8px;
    margin-bottom: 12px;
}

.highlight-metric h4 {
    margin: 0 0 8px 0;
    color: var(--primary);
    font-weight: 600;
}

.highlight-metric p {
    margin: 0;
    font-size: 18px;
    font-weight: 700;
}

/* Status indicators */
.status-indicator {
    display: inline-block;
    width: 12px;
    height: 12px;
    border-radius: 50%;
    margin-right: 8px;
}

.status-active {
    background-color: var(--success);
}

.status-inactive {
    background-color: var(--secondary);
}

.status-warning {
    background-color: var(--warning);
}

.status-error {
    background-color: var(--danger);
}
</style>
"""

# ---------------- UTILITY: NSE Session ----------------
@st.cache_data(ttl=300)
def get_nse_session():
    """
    Create a requests session with common NSE headers to fetch option chain.
    Using a session helps maintain cookies and reduces repeated costs.
    """
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://www.nseindia.com/option-chain"
    })
    # Try a warm-up GET to set cookies (may fail inside some sandboxed envs)
    try:
        s.get("https://www.nseindia.com/option-chain", timeout=8)
    except Exception:
        pass
    return s

# ---------------- FETCH / PARSE ----------------
def fetch_option_chain(symbol, session):
    """
    Fetch option-chain JSON for a symbol. Retries 3 times with backoff.
    Returns JSON or None.
    """
    url = (f"https://www.nseindia.com/api/option-chain-indices?symbol={symbol}"
           if symbol in INDICES else f"https://www.nseindia.com/api/option-chain-equities?symbol={symbol}")
    for attempt in range(3):
        try:
            resp = session.get(url, timeout=30)
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            # backoff with jitter
            time.sleep(random.uniform(1.0, 2.5) * (attempt + 1))
    return None

def parse_data(symbol, data):
    """
    Convert NSE option-chain JSON to a DataFrame for the nearest expiry.
    Columns: STRIKE, CALL_OI, CALL_CHNG_IN_OI, CALL_IV, CALL_LTP, PUT_OI, PUT_CHNG_IN_OI, PUT_IV, PUT_LTP
    """
    if not data:
        return pd.DataFrame()
    
    records = data.get("records", {}) or {}
    expiry_dates = records.get("expiryDates", []) or []
    
    if not expiry_dates:
        st.warning(f"No expiry dates found for {symbol}")
        return pd.DataFrame()
    
    # Get the nearest expiry
    expiry = expiry_dates[0]
    rows = []
    
    for item in records.get("data", []) or []:
        if item.get("expiryDate") != expiry:
            continue
            
        ce = item.get("CE") or {}
        pe = item.get("PE") or {}
        
        rows.append({
            "STRIKE": item.get("strikePrice"),
            "CALL_OI": ce.get("openInterest", 0),
            "CALL_CHNG_IN_OI": ce.get("changeinOpenInterest", 0),
            "CALL_IV": ce.get("impliedVolatility", 0),
            "CALL_LTP": ce.get("lastPrice", 0),
            "PUT_OI": pe.get("openInterest", 0),
            "PUT_CHNG_IN_OI": pe.get("changeinOpenInterest", 0),
            "PUT_IV": pe.get("impliedVolatility", 0),
            "PUT_LTP": pe.get("lastPrice", 0)
        })
    
    df = pd.DataFrame(rows)
    if df.empty:
        st.warning(f"No option data found for {symbol} for expiry {expiry}")
        return df
        
    return df.sort_values("STRIKE").reset_index(drop=True)

# ---------------- ENHANCED ANALYTICS ----------------
def calculate_analytics(df, spot_price=None):
    """
    Enhanced analytics:
    - total/ATM PCR
    - true max pain (loss-based)
    - support/resistance by OI buildup
    - IV skew and approx expected 30-day move
    - directional score
    - volatility surface analysis
    - option flow analysis
    Returns a dict with rich metrics and modified df (with extras).
    """
    if df.empty:
        return {}
        
    df = df.copy()
    
    # estimate spot if not provided: strike with smallest difference between call & put OI
    if not spot_price or spot_price == 0:
        df['OI_DIFF'] = (df['CALL_OI'] - df['PUT_OI']).abs()
        spot_price = int(df.loc[df['OI_DIFF'].idxmin(), "STRIKE"])
        
    df['TOTAL_OI'] = df['CALL_OI'] + df['PUT_OI']
    df['OI_RATIO'] = df['PUT_OI'] / (df['CALL_OI'] + 1e-10)  # Avoid division by zero
    df['DELTA_CALL'] = df['CALL_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['DELTA_PUT'] = df['PUT_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['IV_DIFF'] = df['CALL_IV'] - df['PUT_IV']
    df['PRICE_RATIO'] = df['CALL_LTP'] / (df['PUT_LTP'] + 1e-10)
    
    total_call_oi = df['CALL_OI'].sum()
    total_put_oi = df['PUT_OI'].sum()
    pcr = total_put_oi / total_call_oi if total_call_oi > 0 else 0

    # ATM window: pick nearest strike and +/- 3 strikes (adjust for strike step)
    strike_step = int(df['STRIKE'].diff().median() or 50)
    atm_strike = int(df.iloc[(df['STRIKE'] - spot_price).abs().argsort()[:1]]['STRIKE'].values[0])
    window = 3 * strike_step
    window_df = df[(df['STRIKE'] >= atm_strike - window) & (df['STRIKE'] <= atm_strike + window)]
    atm_pcr = window_df['PUT_OI'].sum() / window_df['CALL_OI'].sum() if window_df['CALL_OI'].sum() > 0 else 0

    # true max pain computation (min aggregated option seller loss)
    strikes = df['STRIKE'].values
    losses = []
    for k in strikes:
        call_loss = ((np.clip(k - strikes, 0, None)) * df['CALL_OI']).sum()
        put_loss = ((np.clip(strikes - k, 0, None)) * df['PUT_OI']).sum()
        losses.append(call_loss + put_loss)
    max_pain = int(strikes[np.argmin(losses)]) if len(losses) > 0 else 0

    # support/resistance by OI buildup and current OI
    df['PUT_SCORE'] = df['PUT_OI'] * (1 + df['PUT_CHNG_IN_OI'] / (df['PUT_OI'].replace(0, 1) + 1))
    df['CALL_SCORE'] = df['CALL_OI'] * (1 + df['CALL_CHNG_IN_OI'] / (df['CALL_OI'].replace(0, 1) + 1))
    strongest_support = int(df.loc[df['PUT_SCORE'].idxmax(), 'STRIKE'])
    strongest_resistance = int(df.loc[df['CALL_SCORE'].idxmax(), 'STRIKE'])

    # IV metrics
    avg_call_iv = df['CALL_IV'].mean()
    avg_put_iv = df['PUT_IV'].mean()
    iv_skew = round(avg_call_iv - avg_put_iv, 2)
    iv_atm = window_df[['CALL_IV', 'PUT_IV']].mean().mean() if not window_df.empty else (avg_call_iv + avg_put_iv) / 2
    expected_move_annual = iv_atm / 100.0
    # approximate 30d expected move (sqrt scaling for monthly ~ sqrt(1/12) of annual vol)
    expected_move_30d = round(spot_price * expected_move_annual / np.sqrt(12), 2)
    
    # Volatility surface analysis
    iv_slope_call = np.polyfit(df['STRIKE'], df['CALL_IV'], 1)[0] if len(df) > 1 else 0
    iv_slope_put = np.polyfit(df['STRIKE'], df['PUT_IV'], 1)[0] if len(df) > 1 else 0
    
    # Option flow analysis
    call_flow = (df['CALL_CHNG_IN_OI'] * df['CALL_LTP']).sum()
    put_flow = (df['PUT_CHNG_IN_OI'] * df['PUT_LTP']).sum()
    net_flow = call_flow - put_flow
    flow_ratio = abs(call_flow / put_flow) if put_flow != 0 else float('inf')

    # volume ratio and direction score
    call_vol = df['CALL_CHNG_IN_OI'].sum()
    put_vol = df['PUT_CHNG_IN_OI'].sum()
    vol_ratio = (call_vol + 1) / (put_vol + 1)
    dir_score = 0.0
    dir_score += (pcr - 1.0) * 2.0
    dir_score += (iv_skew / 5.0)
    dir_score += (vol_ratio - 1.0)
    dir_score += (net_flow / max(abs(net_flow), 1)) * 0.5  # Add flow impact
    
    if dir_score > 1.2:
        direction = "Bullish"
    elif dir_score < -1.2:
        direction = "Bearish"
    else:
        direction = "Neutral"

    top_calls = df.nlargest(5, 'CALL_OI')[['STRIKE', 'CALL_OI', 'CALL_IV', 'CALL_CHNG_IN_OI', 'CALL_LTP']]
    top_puts = df.nlargest(5, 'PUT_OI')[['STRIKE', 'PUT_OI', 'PUT_IV', 'PUT_CHNG_IN_OI', 'PUT_LTP']]
    
    # Option clustering for pattern recognition
    if len(df) > 10:
        try:
            # Prepare data for clustering
            cluster_data = df[['STRIKE', 'CALL_OI', 'PUT_OI', 'CALL_IV', 'PUT_IV']].copy()
            cluster_data = (cluster_data - cluster_data.mean()) / cluster_data.std()
            
            # Apply K-means clustering
            kmeans = KMeans(n_clusters=3, random_state=42, n_init=10)
            df['CLUSTER'] = kmeans.fit_predict(cluster_data)
        except:
            df['CLUSTER'] = 0

    # Additional financial analysis metrics
    # VIX-like volatility indicator
    vix_like = (avg_call_iv + avg_put_iv) / 2
    
    # Put-Call parity deviation
    put_call_parity_dev = (df['CALL_LTP'] - df['PUT_LTP'] + df['STRIKE'] - spot_price).abs().mean()
    
    # Gamma exposure estimation
    gamma_exposure = (df['CALL_OI'] + df['PUT_OI']).sum() / 1e6  # In millions
    
    return {
        "df": df,
        "spot_price": spot_price,
        "pcr": round(pcr, 2),
        "pcr_atm": round(atm_pcr, 2),
        "max_pain": max_pain,
        "support": strongest_support,
        "resistance": strongest_resistance,
        "iv_skew": iv_skew,
        "iv_slope_call": round(iv_slope_call, 4),
        "iv_slope_put": round(iv_slope_put, 4),
        "expected_move_30d": expected_move_30d,
        "direction": direction,
        "dir_score": round(dir_score, 2),
        "call_flow": round(call_flow, 2),
        "put_flow": round(put_flow, 2),
        "net_flow": round(net_flow, 2),
        "flow_ratio": round(flow_ratio, 2),
        "top_calls": top_calls,
        "top_puts": top_puts,
        "vix_like": round(vix_like, 2),
        "put_call_parity_dev": round(put_call_parity_dev, 2),
        "gamma_exposure": round(gamma_exposure, 2)
    }

# ---------------- ML: Regression & Classification ----------------
def train_ml_models_regression(df):
    """
    Regression models to predict 'attractive' strikes (explainable).
    Returns models' results dict and recommended top calls/puts by prediction.
    """
    if df.empty or len(df) < 10:
        return {}, [], [], {}
    
    # Create target variable based on OI and price action
    df_ml = df.copy()
    df_ml['TARGET'] = df_ml['CALL_OI'] * df_ml['CALL_LTP'] - df_ml['PUT_OI'] * df_ml['PUT_LTP']
    
    X = df_ml[['CALL_OI', 'PUT_OI', 'CALL_CHNG_IN_OI', 'PUT_CHNG_IN_OI', 'CALL_IV', 'PUT_IV', 'CALL_LTP', 'PUT_LTP']].fillna(0)
    y = df_ml['TARGET']
    
    # Feature importance using Random Forest
    rf = RandomForestRegressor(n_estimators=100, random_state=42)
    rf.fit(X, y)
    feature_importance = dict(zip(X.columns, rf.feature_importances_))
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    scaler = StandardScaler()
    X_train_s = scaler.fit_transform(X_train)
    X_test_s = scaler.transform(X_test)

    models = {
        "Linear": LinearRegression(),
        "Ridge": Ridge(alpha=1.0),
        "Lasso": Lasso(alpha=0.01),
        "RF": RandomForestRegressor(n_estimators=150, random_state=42),
        "GB": GradientBoostingRegressor(n_estimators=150, random_state=42)
    }
    results = {}
    for name, model in models.items():
        model.fit(X_train_s, y_train)
        y_pred = model.predict(X_test_s)
        mae = mean_absolute_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)
        results[name] = {"model": model, "mae": mae, "r2": r2, "scaler": scaler}

    best_name = min(results, key=lambda k: results[k]['mae'])
    best_model = results[best_name]['model']
    best_scaler = results[best_name]['scaler']

    X_full = scaler.transform(X)
    df_ml['ML_PREDICTED_VALUE'] = best_model.predict(X_full)
    
    # Get top calls and puts based on ML predictions
    top_calls = df_ml.nlargest(3, 'ML_PREDICTED_VALUE')['STRIKE'].tolist()
    top_puts = df_ml.nsmallest(3, 'ML_PREDICTED_VALUE')['STRIKE'].tolist()

    return results, top_calls, top_puts, feature_importance

def train_ml_models_classification(df):
    """
    Classification to detect bias from options features.
    Simple labeled by PCR thresholds.
    """
    if df.empty or len(df) < 12:
        return {"RF": 0.0, "LR": 0.0}, "Neutral", {}
    
    d = df.copy()
    d['PCR'] = d['PUT_OI'] / (d['CALL_OI'] + 1)
    d['IVS'] = d['CALL_IV'] - d['PUT_IV']
    d['FLOW_RATIO'] = (d['CALL_CHNG_IN_OI'] * d['CALL_LTP']) / (d['PUT_CHNG_IN_OI'] * d['PUT_LTP'] + 1)
    
    features = ['CALL_OI', 'PUT_OI', 'PCR', 'IVS', 'CALL_CHNG_IN_OI', 'PUT_CHNG_IN_OI', 'FLOW_RATIO']
    d['LABEL'] = np.where(d['PCR'] > 1.2, 'Bullish', np.where(d['PCR'] < 0.8, 'Bearish', 'Neutral'))
    
    # Handle the ValueError by ensuring all classes have at least 2 members
    y_counts = d['LABEL'].value_counts()
    valid_classes = y_counts[y_counts >= 2].index
    d_filtered = d[d['LABEL'].isin(valid_classes)]
    
    if len(d_filtered) < 12 or len(d_filtered['LABEL'].unique()) < 2:
        return {"RF": 0.0, "LR": 0.0}, "Neutral", {}

    X = d_filtered[features].fillna(0)
    y = d_filtered['LABEL']
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
    
    rf = RandomForestClassifier(n_estimators=150, random_state=42)
    rf.fit(X_train, y_train)
    acc_rf = accuracy_score(y_test, rf.predict(X_test))
    feature_importance = dict(zip(features, rf.feature_importances_))
    
    lr = LogisticRegression(max_iter=500, class_weight='balanced')
    lr.fit(X_train, y_train)
    acc_lr = accuracy_score(y_test, lr.predict(X_test))
    
    # consensus label from RF predictions on the set (mode)
    preds = rf.predict(X)
    consensus = pd.Series(preds).mode().iloc[0] if len(preds) > 0 else "Neutral"
    
    # Confusion matrix and classification report
    y_pred = rf.predict(X_test)
    cm = confusion_matrix(y_test, y_pred, labels=['Bullish', 'Neutral', 'Bearish'])
    cr = classification_report(y_test, y_pred, output_dict=True, zero_division=0)
    
    return {"RF": round(acc_rf, 3), "LR": round(acc_lr, 3)}, consensus, {
        "feature_importance": feature_importance,
        "confusion_matrix": cm,
        "classification_report": cr
    }

# ---------------- ADVANCED CHARTS ----------------
def create_oi_chart(df, line_shape='spline'):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['CALL_OI'], mode='lines', name='Call OI', 
                             line=dict(shape=line_shape, width=3, color='#FF6B6B')))
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['PUT_OI'], mode='lines', name='Put OI', 
                             line=dict(shape=line_shape, width=3, color='#4ECDC4')))
    fig.update_layout(
        title="Open Interest Distribution",
        xaxis_title="Strike Price",
        yaxis_title="Open Interest",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_sentiment_chart(df, line_shape='spline'):
    df_local = df.copy()
    df_local['SENT'] = df_local['CALL_OI'] - df_local['PUT_OI']
    fig = go.Figure([go.Bar(x=df_local['STRIKE'], y=df_local['SENT'], 
                           marker_color=np.where(df_local['SENT'] > 0, '#FF6B6B', '#4ECDC4'))])
    fig.update_layout(
        title="Sentiment (Call OI - Put OI)",
        xaxis_title="Strike Price",
        yaxis_title="Call-Put OI Difference",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_iv_comparison_chart(df, line_shape='spline'):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['CALL_IV'], name='Call IV', 
                             line=dict(shape=line_shape, width=3, color='#FF6B6B')))
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['PUT_IV'], name='Put IV', 
                             line=dict(shape=line_shape, width=3, color='#4ECDC4')))
    fig.update_layout(
        title="Implied Volatility (Call vs Put)",
        xaxis_title="Strike Price",
        yaxis_title="Implied Volatility (%)",
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_volatility_surface_chart(df):
    fig = make_subplots(rows=1, cols=2, subplot_titles=("Call IV Surface", "Put IV Surface"))
    
    # Call IV surface
    fig.add_trace(
        go.Scatter(x=df['STRIKE'], y=df['CALL_IV'], mode='lines+markers', 
                  name='Call IV', line=dict(width=3, color='#FF6B6B')),
        row=1, col=1
    )
    
    # Put IV surface
    fig.add_trace(
        go.Scatter(x=df['STRIKE'], y=df['PUT_IV'], mode='lines+markers', 
                  name='Put IV', line=dict(width=3, color='#4ECDC4')),
        row=1, col=2
    )
    
    fig.update_xaxes(title_text="Strike Price", row=1, col=1)
    fig.update_xaxes(title_text="Strike Price", row=1, col=2)
    fig.update_yaxes(title_text="Implied Volatility (%)", row=1, col=1)
    fig.update_yaxes(title_text="Implied Volatility (%)", row=1, col=2)
    
    fig.update_layout(
        title="Volatility Surface Analysis",
        height=400,
        margin=dict(t=80, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    
    return fig

def create_ml_prediction_chart(df, analytics, top_calls, top_puts, line_shape='spline'):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['TOTAL_OI'], mode='lines+markers', 
                             name='Total OI', line=dict(shape=line_shape, width=3, color='#45B7D1')))
    
    if 'ML_PREDICTED_VALUE' in df.columns:
        fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['ML_PREDICTED_VALUE'], mode='markers', 
                                name='ML Predicted', marker=dict(size=8, color='#FBE555')))
    
    fig.add_vline(x=analytics.get('max_pain', None), line_dash='dash', line_color='#964B00', 
                  annotation_text='Max Pain', annotation_position="top right")
    
    for s in top_calls:
        fig.add_vline(x=s, line_dash='dot', line_color='green', annotation_text=f'Call {s}')
    
    for s in top_puts:
        fig.add_vline(x=s, line_dash='dot', line_color='red', annotation_text=f'Put {s}')
    
    fig.update_layout(
        title="ML Predicted Strikes & OI",
        xaxis_title="Strike Price",
        yaxis_title="Total OI / ML Prediction",
        height=450,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    return fig

def create_model_performance_chart(ml_results):
    if not ml_results:
        return go.Figure()
    
    models = list(ml_results.keys())
    mae = [ml_results[m]['mae'] for m in models]
    r2 = [ml_results[m]['r2'] for m in models]
    
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    
    # Bar for MAE
    fig.add_trace(
        go.Bar(
            x=models,
            y=mae,
            name='MAE',
            marker_color='orange'
        ),
        secondary_y=False
    )
    
    # Line for R2
    fig.add_trace(
        go.Scatter(
            x=models,
            y=r2,
            name='R²',
            mode='lines+markers',
            line=dict(color='blue', width=2),
            marker=dict(size=8)
        ),
        secondary_y=True
    )
    
    # Layout with dual y-axis
    fig.update_layout(
        title="Regression Model Performance",
        xaxis_title="Model",
        yaxis=dict(title="MAE", side='left', showgrid=False),
        yaxis2=dict(title="R²", side='right', showgrid=False, range=[-1, 1]),
        height=400,
        margin=dict(t=50, b=50, l=50, r=50),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12)
    )
    
    return fig

# ---------------- PROFESSIONAL EXCEL EXPORT ----------------
def create_excel_export(df, analytics, symbol, ml_results, top_calls, top_puts, session):
    """
    Create a professional Excel workbook with multiple sheets, formatting, and charts.
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Data Sheet
    ws_data = wb.create_sheet("Option Chain Data")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    # Format data sheet
    for cell in ws_data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=1, max_col=ws_data.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Analytics Sheet
    ws_analytics = wb.create_sheet("Analytics")
    analytics_data = [
        ["Metric", "Value"],
        ["Symbol", symbol],
        ["Spot Price", analytics.get("spot_price", "N/A")],
        ["PCR (Total)", analytics.get("pcr", "N/A")],
        ["PCR (ATM)", analytics.get("pcr_atm", "N/A")],
        ["Max Pain", analytics.get("max_pain", "N/A")],
        ["Support", analytics.get("support", "N/A")],
        ["Resistance", analytics.get("resistance", "N/A")],
        ["IV Skew", analytics.get("iv_skew", "N/A")],
        ["IV Slope (Call)", analytics.get("iv_slope_call", "N/A")],
        ["IV Slope (Put)", analytics.get("iv_slope_put", "N/A")],
        ["Expected 30D Move", analytics.get("expected_move_30d", "N/A")],
        ["Direction", analytics.get("direction", "N/A")],
        ["Direction Score", analytics.get("dir_score", "N/A")],
        ["Call Flow", analytics.get("call_flow", "N/A")],
        ["Put Flow", analytics.get("put_flow", "N/A")],
        ["Net Flow", analytics.get("net_flow", "N/A")],
        ["Flow Ratio", analytics.get("flow_ratio", "N/A")],
        ["VIX-like Indicator", analytics.get("vix_like", "N/A")],
        ["Put-Call Parity Dev", analytics.get("put_call_parity_dev", "N/A")],
        ["Gamma Exposure", analytics.get("gamma_exposure", "N/A")]
    ]
    
    for row in analytics_data:
        ws_analytics.append(row)
    
    # Format analytics sheet
    for cell in ws_analytics[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row in ws_analytics.iter_rows(min_row=2, max_row=ws_analytics.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")
    
    # ML Results Sheet
    if ml_results:
        ws_ml = wb.create_sheet("ML Results")
        ml_data = [["Model", "MAE", "R²"]]
        for model_name, result in ml_results.items():
            ml_data.append([model_name, result.get("mae", "N/A"), result.get("r2", "N/A")])
        
        ml_data.append([])
        ml_data.append(["Top Calls (ML)", "Top Puts (ML)"])
        for i in range(max(len(top_calls), len(top_puts))):
            call_val = top_calls[i] if i < len(top_calls) else ""
            put_val = top_puts[i] if i < len(top_puts) else ""
            ml_data.append([call_val, put_val])
        
        for row in ml_data:
            ws_ml.append(row)
        
        # Format ML sheet
        for cell in ws_ml[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row in ws_ml.iter_rows(min_row=2, max_row=ws_ml.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    
    # Top OI Sheet
    ws_top = wb.create_sheet("Top OI")
    top_calls_data = analytics.get("top_calls", pd.DataFrame())
    top_puts_data = analytics.get("top_puts", pd.DataFrame())
    
    if not top_calls_data.empty:
        ws_top.append(["Top Calls by OI"])
        for r in dataframe_to_rows(top_calls_data, index=False, header=True):
            ws_top.append(r)
        ws_top.append([])
    
    if not top_puts_data.empty:
        ws_top.append(["Top Puts by OI"])
        for r in dataframe_to_rows(top_puts_data, index=False, header=True):
            ws_top.append(r)
    
    # Format top OI sheet
    for row in ws_top.iter_rows():
        for cell in row:
            if cell.value and cell.value in ["Top Calls by OI", "Top Puts by OI"]:
                cell.font = Font(bold=True, size=14)
    
    # Add charts
    try:
        # OI Chart
        oi_fig = create_oi_chart(df)
        oi_img_bytes = pio.to_image(oi_fig, format='png', width=800, height=400)
        oi_img = XLImage(BytesIO(oi_img_bytes))
        oi_img.width = 600
        oi_img.height = 300
        
        ws_charts = wb.create_sheet("Charts")
        ws_charts.add_image(oi_img, "A1")
        
        # IV Chart
        iv_fig = create_iv_comparison_chart(df)
        iv_img_bytes = pio.to_image(iv_fig, format='png', width=800, height=400)
        iv_img = XLImage(BytesIO(iv_img_bytes))
        iv_img.width = 600
        iv_img.height = 300
        ws_charts.add_image(iv_img, "A30")
    except Exception as e:
        print(f"Could not add charts to Excel: {e}")
    
    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

# ---------------- STREAMLIT UI ----------------
def main():
    # Inject CSS
    st.markdown(PRO_CSS, unsafe_allow_html=True)
    
    # Header
    st.markdown("""
    <div class="header">
        <div>
            <h1 class="app-title">NSE Option Chain Analyzer</h1>
            <p class="app-sub">Advanced analytics, ML insights, and professional reporting for NSE options</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize session state
    if 'symbols' not in st.session_state:
        st.session_state.symbols = ["NIFTY", "BANKNIFTY"]
    if 'watchlist' not in st.session_state:
        st.session_state.watchlist = []
    if 'session' not in st.session_state:
        st.session_state.session = get_nse_session()
    if 'last_update' not in st.session_state:
        st.session_state.last_update = datetime.now()
    if 'auto_refresh' not in st.session_state:
        st.session_state.auto_refresh = False
    
    # Sidebar
    with st.sidebar:
        st.markdown("""
        <div style="padding: 10px; background: #f0f7ff; border-radius: 8px; margin-bottom: 20px;">
            <h3 style="margin: 0; color: #0b69ff;">Settings</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Symbol selection
        symbol_input = st.text_input("Add Symbol (e.g., RELIANCE, INFY):", value="", key="symbol_input")
        if st.button("Add Symbol") and symbol_input:
            if symbol_input.upper() not in st.session_state.symbols:
                st.session_state.symbols.append(symbol_input.upper())
        
        # Index selection
        selected_indices = st.multiselect(
            "Select Indices:",
            options=INDICES,
            default=["NIFTY", "BANKNIFTY"],
            key="selected_indices"
        )
        
        # Update symbols list
        st.session_state.symbols = selected_indices + [s for s in st.session_state.symbols if s not in INDICES]
        
        # Remove symbol
        if st.session_state.symbols:
            to_remove = st.selectbox("Remove Symbol:", options=st.session_state.symbols, key="remove_select")
            if st.button("Remove Selected"):
                st.session_state.symbols.remove(to_remove)
        
        # Watchlist
        st.markdown("---")
        st.markdown("""
        <div style="padding: 10px; background: #f0f7ff; border-radius: 8px; margin-bottom: 20px;">
            <h3 style="margin: 0; color: #0b69ff;">Watchlist</h3>
        </div>
        """, unsafe_allow_html=True)
        
        watch_input = st.text_input("Add to Watchlist:", value="", key="watch_input")
        if st.button("Add to Watchlist") and watch_input:
            if watch_input.upper() not in st.session_state.watchlist:
                st.session_state.watchlist.append(watch_input.upper())
        
        if st.session_state.watchlist:
            to_remove_watch = st.selectbox("Remove from Watchlist:", options=st.session_state.watchlist, key="remove_watch")
            if st.button("Remove from Watchlist"):
                st.session_state.watchlist.remove(to_remove_watch)
        
        # Auto-refresh
        st.markdown("---")
        auto_refresh = st.checkbox("Auto-refresh (every 60s)", value=st.session_state.auto_refresh, key="auto_refresh")
        if auto_refresh != st.session_state.auto_refresh:
            st.session_state.auto_refresh = auto_refresh
        
        if st.button("Refresh Data", key="manual_refresh"):
            st.session_state.last_update = datetime.now()
            st.rerun()
        
        # Display last update time
        st.markdown(f"**Last Update:** {st.session_state.last_update.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Status indicators
        st.markdown("---")
        st.markdown("""
        <div style="padding: 10px; background: #f0f7ff; border-radius: 8px; margin-bottom: 20px;">
            <h3 style="margin: 0; color: #0b69ff;">Status</h3>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            <div style="display: flex; align-items: center;">
                <span class="status-indicator status-active"></span>
                <span>Data Connection</span>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown("""
            <div style="display: flex; align-items: center;">
                <span class="status-indicator status-active"></span>
                <span>ML Models</span>
            </div>
            """, unsafe_allow_html=True)
    
    # Main content with extra tab
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Option Chain", 
        "Analytics & ML", 
        "Strategy Signals", 
        "Backtesting",
        "Strategies & Greeks"
    ])
    
    with tab1:
        st.markdown("""
        <div class="section">
            <h2 class="section-title">Option Chain Data</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if not st.session_state.symbols:
            st.info("Please add at least one symbol from the sidebar.")
        else:
            # Fetch data for all symbols
            with st.spinner("Fetching option chain data..."):
                all_data = {}
                with concurrent.futures.ThreadPoolExecutor(max_workers=THREAD_WORKERS) as executor:
                    future_to_symbol = {
                        executor.submit(fetch_option_chain, symbol, st.session_state.session): symbol
                        for symbol in st.session_state.symbols
                    }
                    
                    for future in concurrent.futures.as_completed(future_to_symbol):
                        symbol = future_to_symbol[future]
                        try:
                            data = future.result()
                            all_data[symbol] = data
                        except Exception as e:
                            st.error(f"Error fetching data for {symbol}: {str(e)}")
                
                # Process and display data
                for symbol in st.session_state.symbols:
                    if symbol not in all_data or not all_data[symbol]:
                        st.warning(f"No data for {symbol}")
                        continue
                    
                    df = parse_data(symbol, all_data[symbol])
                    if df.empty:
                        st.warning(f"Empty DataFrame for {symbol}")
                        continue
                    
                    analytics = calculate_analytics(df)
                    
                    # Display metrics
                    st.markdown(f"### {symbol}")
                    
                    col1, col2, col3, col4, col5 = st.columns(5)
                    with col1:
                        st.markdown(f"""
                        <div class="kpi">
                            <div class="value">{analytics.get('spot_price', 'N/A')}</div>
                            <div class="label">Spot Price</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col2:
                        pcr_class = "bullish" if analytics.get('pcr', 0) > 1 else "bearish" if analytics.get('pcr', 0) < 1 else "neutral"
                        st.markdown(f"""
                        <div class="kpi">
                            <div class="value {pcr_class}">{analytics.get('pcr', 'N/A')}</div>
                            <div class="label">PCR</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col3:
                        st.markdown(f"""
                        <div class="kpi">
                            <div class="value">{analytics.get('max_pain', 'N/A')}</div>
                            <div class="label">Max Pain</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col4:
                        dir_class = "bullish" if analytics.get('direction') == "Bullish" else "bearish" if analytics.get('direction') == "Bearish" else "neutral"
                        st.markdown(f"""
                        <div class="kpi">
                            <div class="value {dir_class}">{analytics.get('direction', 'N/A')}</div>
                            <div class="label">Direction</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col5:
                        st.markdown(f"""
                        <div class="kpi">
                            <div class="value">{analytics.get('expected_move_30d', 'N/A')}</div>
                            <div class="label">Exp. 30D Move</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Display data table
                    st.dataframe(df.style.format({
                        'CALL_OI': '{:,.0f}',
                        'CALL_CHNG_IN_OI': '{:,.0f}',
                        'CALL_IV': '{:.2f}',
                        'CALL_LTP': '{:.2f}',
                        'PUT_OI': '{:,.0f}',
                        'PUT_CHNG_IN_OI': '{:,.0f}',
                        'PUT_IV': '{:.2f}',
                        'PUT_LTP': '{:.2f}'
                    }), height=300, width='stretch')
                    
                    # Export button
                    excel_buffer = create_excel_export(
                        df, analytics, symbol, {}, [], [], st.session_state.session
                    )
                    st.download_button(
                        label=f"Download {symbol} Data as Excel",
                        data=excel_buffer,
                        file_name=f"{symbol}_option_chain_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{symbol}"
                    )
    
    with tab2:
        st.markdown("""
        <div class="section">
            <h2 class="section-title">Advanced Analytics & Machine Learning</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if not st.session_state.symbols:
            st.info("Please add at least one symbol from the sidebar.")
        else:
            selected_symbol = st.selectbox("Select Symbol for Analysis:", options=st.session_state.symbols, key="analysis_symbol")
            
            if selected_symbol:
                with st.spinner(f"Fetching and analyzing data for {selected_symbol}..."):
                    data = fetch_option_chain(selected_symbol, st.session_state.session)
                    if not data:
                        st.error(f"Failed to fetch data for {selected_symbol}")
                    else:
                        df = parse_data(selected_symbol, data)
                        if df.empty:
                            st.warning(f"Empty DataFrame for {selected_symbol}")
                        else:
                            analytics = calculate_analytics(df)
                            
                            # Display advanced analytics
                            st.markdown(f"### Advanced Analytics for {selected_symbol}")
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.markdown(f"""
                                <div class="highlight-metric">
                                    <h4>Volatility Analysis</h4>
                                    <p>IV Skew: {analytics.get('iv_skew', 'N/A')}</p>
                                    <p>VIX-like: {analytics.get('vix_like', 'N/A')}%</p>
                                    <p>Call IV Slope: {analytics.get('iv_slope_call', 'N/A')}</p>
                                    <p>Put IV Slope: {analytics.get('iv_slope_put', 'N/A')}</p>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            with col2:
                                st.markdown(f"""
                                <div class="highlight-metric">
                                    <h4>Flow Analysis</h4>
                                    <p>Call Flow: ₹{analytics.get('call_flow', 'N/A'):,.0f}</p>
                                    <p>Put Flow: ₹{analytics.get('put_flow', 'N/A'):,.0f}</p>
                                    <p>Net Flow: ₹{analytics.get('net_flow', 'N/A'):,.0f}</p>
                                    <p>Flow Ratio: {analytics.get('flow_ratio', 'N/A'):.2f}</p>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            with col3:
                                st.markdown(f"""
                                <div class="highlight-metric">
                                    <h4>Risk Metrics</h4>
                                    <p>Gamma Exposure: {analytics.get('gamma_exposure', 'N/A')}M</p>
                                    <p>Put-Call Parity Dev: {analytics.get('put_call_parity_dev', 'N/A'):.2f}</p>
                                    <p>Support: {analytics.get('support', 'N/A')}</p>
                                    <p>Resistance: {analytics.get('resistance', 'N/A')}</p>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            # Charts
                            st.plotly_chart(create_oi_chart(analytics['df']), width='stretch')
                            st.plotly_chart(create_iv_comparison_chart(analytics['df']), width='stretch')
                            st.plotly_chart(create_volatility_surface_chart(analytics['df']), width='stretch')
                            
                            # ML Analysis
                            st.markdown("### Machine Learning Insights")
                            
                            ml_results, top_calls, top_puts, feature_importance = train_ml_models_regression(analytics['df'])
                            if ml_results:
                                st.plotly_chart(create_ml_prediction_chart(analytics['df'], analytics, top_calls, top_puts), width='stretch')
                                st.plotly_chart(create_model_performance_chart(ml_results), width='stretch')
                                
                                st.markdown("#### Top ML Recommendations")
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Top Calls (ML):**")
                                    for strike in top_calls:
                                        st.write(f"- {strike}")
                                with col2:
                                    st.markdown("**Top Puts (ML):**")
                                    for strike in top_puts:
                                        st.write(f"- {strike}")
                                
                                st.markdown("#### Feature Importance")
                                feature_df = pd.DataFrame(list(feature_importance.items()), columns=['Feature', 'Importance'])
                                st.dataframe(feature_df.sort_values('Importance', ascending=False), width='stretch')
                            else:
                                st.info("Not enough data for ML analysis")
    
    with tab3:
        st.markdown("""
        <div class="section">
            <h2 class="section-title">Strategy Signals</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if not st.session_state.symbols:
            st.info("Please add at least one symbol from the sidebar.")
        else:
            selected_symbol = st.selectbox("Select Symbol:", options=st.session_state.symbols, key="strategy_symbol")
            
            if selected_symbol:
                with st.spinner(f"Generating strategy signals for {selected_symbol}..."):
                    data = fetch_option_chain(selected_symbol, st.session_state.session)
                    if not data:
                        st.error(f"Failed to fetch data for {selected_symbol}")
                    else:
                        df = parse_data(selected_symbol, data)
                        if df.empty:
                            st.warning(f"Empty DataFrame for {selected_symbol}")
                        else:
                            analytics = calculate_analytics(df)
                            
                            # Generate strategy signals
                            signals = []
                            
                            # PCR-based signal
                            if analytics.get('pcr', 0) > 1.5:
                                signals.append(("Bullish PCR", "High PCR indicates bullish sentiment", "BULLISH", "high"))
                            elif analytics.get('pcr', 0) < 0.7:
                                signals.append(("Bearish PCR", "Low PCR indicates bearish sentiment", "BEARISH", "high"))
                            
                            # IV Skew signal
                            if analytics.get('iv_skew', 0) > 2:
                                signals.append(("Call Skew", "Call IV higher than Put IV, suggests bullish expectations", "BULLISH", "medium"))
                            elif analytics.get('iv_skew', 0) < -2:
                                signals.append(("Put Skew", "Put IV higher than Call IV, suggests bearish expectations", "BEARISH", "medium"))
                            
                            # Flow-based signal
                            if analytics.get('net_flow', 0) > 1000000:
                                signals.append(("Call Buying Pressure", "Significant net call flow indicates bullish positioning", "BULLISH", "medium"))
                            elif analytics.get('net_flow', 0) < -1000000:
                                signals.append(("Put Buying Pressure", "Significant net put flow indicates bearish positioning", "BEARISH", "medium"))
                            
                            # Max Pain signal
                            spot = analytics.get('spot_price', 0)
                            max_pain = analytics.get('max_pain', 0)
                            if spot > max_pain * 1.02:
                                signals.append(("Above Max Pain", "Price above max pain, potential resistance", "BEARISH", "low"))
                            elif spot < max_pain * 0.98:
                                signals.append(("Below Max Pain", "Price below max pain, potential support", "BULLISH", "low"))
                            
                            # Display signals
                            if not signals:
                                st.info("No strong signals detected for this symbol")
                            else:
                                for signal, description, direction, strength in signals:
                                    direction_color = "green" if direction == "BULLISH" else "red" if direction == "BEARISH" else "gray"
                                    strength_color = "red" if strength == "high" else "orange" if strength == "medium" else "gray"
                                    
                                    st.markdown(f"""
                                    <div style="padding: 12px; border-radius: 8px; border-left: 4px solid {direction_color}; background: #f9f9f9; margin-bottom: 10px;">
                                        <div style="display: flex; justify-content: space-between; align-items: center;">
                                            <h4 style="margin: 0;">{signal}</h4>
                                            <span style="background: {strength_color}; color: white; padding: 4px 8px; border-radius: 4px; font-size: 12px;">{strength.upper()}</span>
                                        </div>
                                        <p style="margin: 8px 0 0 0; color: #666;">{description}</p>
                                        <p style="margin: 4px 0 0 0; color: {direction_color}; font-weight: bold;">{direction}</p>
                                    </div>
                                    """, unsafe_allow_html=True)
    
    with tab4:
        st.markdown("""
        <div class="section">
            <h2 class="section-title">Backtesting & Historical Analysis</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if not st.session_state.watchlist:
            st.info("Add symbols to your watchlist in the sidebar to enable backtesting.")
        else:
            selected_symbol = st.selectbox("Select Symbol for Backtesting:", options=st.session_state.watchlist, key="backtest_symbol")
            period = st.selectbox("Select Period:", options=["1mo", "3mo", "6mo", "1y", "2y"], index=2, key="backtest_period")
            
            if selected_symbol and period:
                with st.spinner(f"Fetching historical data for {selected_symbol}..."):
                    try:
                        # Fetch historical data using yfinance
                        ticker = yf.Ticker(f"{selected_symbol}.NS")
                        hist = ticker.history(period=period)
                        
                        if hist.empty:
                            st.error(f"No historical data found for {selected_symbol}")
                        else:
                            # Calculate technical indicators
                            hist['SMA_20'] = hist['Close'].rolling(window=20).mean()
                            hist['SMA_50'] = hist['Close'].rolling(window=50).mean()
                            hist['RSI'] = calculate_rsi(hist['Close'])
                            hist['MACD'], hist['MACD_Signal'], hist['MACD_Hist'] = calculate_macd(hist['Close'])
                            
                            # Display price chart with indicators
                            fig = go.Figure()
                            fig.add_trace(go.Candlestick(
                                x=hist.index,
                                open=hist['Open'],
                                high=hist['High'],
                                low=hist['Low'],
                                close=hist['Close'],
                                name='OHLC'
                            ))
                            fig.add_trace(go.Scatter(x=hist.index, y=hist['SMA_20'], name='SMA 20', line=dict(color='orange', width=1)))
                            fig.add_trace(go.Scatter(x=hist.index, y=hist['SMA_50'], name='SMA 50', line=dict(color='blue', width=1)))
                            
                            fig.update_layout(
                                title=f"{selected_symbol} Price Chart with Moving Averages",
                                yaxis_title="Price (₹)",
                                height=500,
                                xaxis_rangeslider_visible=False
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Display technical indicators
                            fig_tech = make_subplots(rows=2, cols=1, subplot_titles=["RSI", "MACD"])
                            
                            # RSI
                            fig_tech.add_trace(
                                go.Scatter(x=hist.index, y=hist['RSI'], name='RSI', line=dict(color='purple', width=1)),
                                row=1, col=1
                            )
                            fig_tech.add_hline(y=70, line_dash="dash", line_color="red", row=1, col=1)
                            fig_tech.add_hline(y=30, line_dash="dash", line_color="green", row=1, col=1)
                            fig_tech.update_yaxes(title_text="RSI", range=[0, 100], row=1, col=1)
                            
                            # MACD
                            fig_tech.add_trace(
                                go.Scatter(x=hist.index, y=hist['MACD'], name='MACD', line=dict(color='blue', width=1)),
                                row=2, col=1
                            )
                            fig_tech.add_trace(
                                go.Scatter(x=hist.index, y=hist['MACD_Signal'], name='Signal', line=dict(color='red', width=1)),
                                row=2, col=1
                            )
                            fig_tech.add_trace(
                                go.Bar(x=hist.index, y=hist['MACD_Hist'], name='Histogram', marker_color=np.where(hist['MACD_Hist'] < 0, 'red', 'green')),
                                row=2, col=1
                            )
                            fig_tech.update_yaxes(title_text="MACD", row=2, col=1)
                            
                            fig_tech.update_layout(
                                height=500,
                                showlegend=True
                            )
                            
                            st.plotly_chart(fig_tech, use_container_width=True)
                            
                            # Performance statistics
                            returns = hist['Close'].pct_change().dropna()
                            cumulative_returns = (1 + returns).cumprod() - 1
                            
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Total Return", f"{cumulative_returns.iloc[-1] * 100:.2f}%")
                            with col2:
                                st.metric("Annual Volatility", f"{returns.std() * np.sqrt(252) * 100:.2f}%")
                            with col3:
                                st.metric("Sharpe Ratio", f"{(returns.mean() * 252) / (returns.std() * np.sqrt(252)):.2f}" if returns.std() > 0 else "N/A")
                            with col4:
                                st.metric("Max Drawdown", f"{calculate_max_drawdown(hist['Close']):.2f}%")
                            
                    except Exception as e:
                        st.error(f"Error fetching historical data: {str(e)}")
    
    with tab5:
        st.subheader("📈 Strategy Payoff Simulator + Greeks")

        # Inputs
        spot = st.number_input("Spot Price", value=20000)
        r = st.number_input("Risk-free Rate (%)", value=6.0) / 100
        sigma = st.number_input("Implied Volatility (%)", value=20.0) / 100
        T_days = st.number_input("Days to Expiry", value=30)
        T = T_days/365

        # Greeks
        K = st.number_input("Strike Price (for Greeks)", value=20000)
        option_type = st.radio("Option Type", ["call","put"])
        greeks = option_greeks(spot, K, T, r, sigma, option_type)
        st.write("**Option Greeks:**", greeks)

        # Strategy Selector
        strategy_choice = st.selectbox("Select Strategy", 
                                       ["Long Straddle","Short Straddle","Bull Call Spread","Iron Condor"])

        if strategy_choice == "Long Straddle":
            strategy = [
                {"type":"call","K":spot,"premium":200,"pos":"long"},
                {"type":"put","K":spot,"premium":180,"pos":"long"}
            ]
        elif strategy_choice == "Short Straddle":
            strategy = [
                {"type":"call","K":spot,"premium":200,"pos":"short"},
                {"type":"put","K":spot,"premium":180,"pos":"short"}
            ]
        elif strategy_choice == "Bull Call Spread":
            strategy = [
                {"type":"call","K":spot,"premium":200,"pos":"long"},
                {"type":"call","K":spot+200,"premium":120,"pos":"short"}
            ]
        elif strategy_choice == "Iron Condor":
            strategy = [
                {"type":"call","K":spot+200,"premium":120,"pos":"short"},
                {"type":"call","K":spot+400,"premium":60,"pos":"long"},
                {"type":"put","K":spot-200,"premium":100,"pos":"short"},
                {"type":"put","K":spot-400,"premium":50,"pos":"long"}
            ]

        # Plot Strategy
        plot_strategy(strategy_choice, strategy, spot, [spot-1000, spot+1000])

        # Show Summary Table
        st.markdown("### 📊 Strategy Summary")
        summary = strategy_summary(strategy, spot, [spot-1000, spot+1000])
        st.write(summary)
        
    # Auto-refresh logic
    if st.session_state.auto_refresh:
        now = datetime.now()
        if (now - st.session_state.last_update).seconds >= 600:
            st.session_state.last_update = now
            st.rerun()

# ---------------- TECHNICAL INDICATORS FOR BACKTESTING ----------------
def calculate_rsi(prices, period=14):
    deltas = np.diff(prices)
    seed = deltas[:period+1]
    up = seed[seed >= 0].sum() / period
    down = -seed[seed < 0].sum() / period
    rs = up / down
    rsi = np.zeros_like(prices)
    rsi[:period] = 100. - 100. / (1. + rs)
    
    for i in range(period, len(prices)):
        delta = deltas[i - 1]
        if delta > 0:
            up_val = delta
            down_val = 0.
        else:
            up_val = 0.
            down_val = -delta
        
        up = (up * (period - 1) + up_val) / period
        down = (down * (period - 1) + down_val) / period
        rs = up / down
        rsi[i] = 100. - 100. / (1. + rs)
    
    return rsi

def calculate_macd(prices, fast=12, slow=26, signal=9):
    exp1 = prices.ewm(span=fast, adjust=False).mean()
    exp2 = prices.ewm(span=slow, adjust=False).mean()
    macd = exp1 - exp2
    macd_signal = macd.ewm(span=signal, adjust=False).mean()
    macd_hist = macd - macd_signal
    return macd, macd_signal, macd_hist

def calculate_max_drawdown(prices):
    peak = prices.expanding(min_periods=1).max()
    drawdown = (prices - peak) / peak
    return drawdown.min() * 100


if __name__ == "__main__":
    main()